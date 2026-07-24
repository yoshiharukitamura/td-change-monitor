from __future__ import annotations

import os
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Protocol

import yaml
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from td_change_monitor.config import (
    MonitorStatus,
    ResourceTargetsConfig,
    SavedQueryTarget,
    WorkflowProjectTarget,
)
from td_change_monitor.errors import ExternalApiError
from td_change_monitor.models import (
    SavedQuerySummary,
    WorkflowDefinitionSummary,
    WorkflowScheduleDetail,
)

_WORKFLOW_SHEET = "登録Workflow一覧"
_WORKFLOW_SCHEDULE_SHEET = "Workflowスケジュール一覧"
_SAVED_QUERY_SHEET = "登録クエリ一覧"
_ACTIVE_VALUE = "利用中"


class WorkflowTargetLookupProtocol(Protocol):
    """Workflow対象のproject ID照合に必要なAPI境界を表す。"""

    async def fetch_workflow_page_by_project_name(
        self,
        project_name: str,
        *,
        last_id: int = 0,
        count: int = 3,
    ) -> tuple[WorkflowDefinitionSummary, ...]:
        """project名で検索したWorkflow一覧1ページを返す。"""
        ...

    async def fetch_schedule(self, schedule_id: str) -> WorkflowScheduleDetail:
        """schedule IDから所属projectを含む現在scheduleを返す。"""
        ...


class SavedQueryTargetLookupProtocol(Protocol):
    """登録クエリ対象のQuery ID照合に必要なAPI境界を表す。"""

    async def fetch_queries(self) -> tuple[SavedQuerySummary, ...]:
        """Query IDを含む登録クエリ一覧を全件返す。"""
        ...


@dataclass(frozen=True)
class TargetImportSummary:
    """対象マスター生成時の照合件数を保持する。"""

    workflow_project_count: int
    workflow_needs_review_count: int
    saved_query_count: int
    saved_query_needs_review_count: int

    def as_dict(self) -> dict[str, int]:
        """ログ出力可能な件数辞書へ変換する。

        引数:
            なし。
        戻り値:
            リソース別の総数とneeds_review件数。
        """
        return {
            "workflow_project_count": self.workflow_project_count,
            "workflow_needs_review_count": self.workflow_needs_review_count,
            "saved_query_count": self.saved_query_count,
            "saved_query_needs_review_count": self.saved_query_needs_review_count,
        }


@dataclass(frozen=True)
class _WorkflowInventory:
    """Excelから抽出したproject別の利用中Workflowとscheduleを保持する。"""

    workflows: dict[str, tuple[str, ...]]
    schedule_ids: dict[str, tuple[str, ...]]


@dataclass(frozen=True)
class _SavedQueryInventoryItem:
    """Excelの利用中登録クエリ1行に含まれる初回照合キーを保持する。"""

    query_name: str
    database: str
    owner: str


async def build_resource_targets_from_workbook(
    workbook_path: Path,
    *,
    workflow_client: WorkflowTargetLookupProtocol,
    saved_query_client: SavedQueryTargetLookupProtocol,
) -> tuple[ResourceTargetsConfig, TargetImportSummary]:
    """棚卸しExcelの利用中行を実APIの安定IDと照合して対象マスターを作る。

    引数:
        workbook_path: 指定されたTD利用状況棚卸しExcel。
        workflow_client: project名からproject IDを確認するWorkflow APIクライアント。
        saved_query_client: Query ID付き一覧を取得する登録クエリAPIクライアント。
    戻り値:
        一意照合済みまたはneeds_reviewの対象マスターと件数要約。
    """
    workflow_inventory, saved_query_inventory = _read_workbook(workbook_path)
    workflow_targets: list[WorkflowProjectTarget] = []
    project_names = sorted(
        set(workflow_inventory.workflows) | set(workflow_inventory.schedule_ids)
    )
    for project_name in project_names:
        page = await workflow_client.fetch_workflow_page_by_project_name(project_name)
        project_ids = {
            workflow.project.project_id
            for workflow in page
            if workflow.project.project_name == project_name
        }
        # project名検索で見つからなくても、棚卸しにschedule IDがあれば、
        # 確認済みschedule APIの所属project IDを第二の照合根拠として使用する。
        if not project_ids:
            for schedule_id in workflow_inventory.schedule_ids.get(project_name, ()):
                try:
                    schedule = await workflow_client.fetch_schedule(schedule_id)
                except ExternalApiError as exc:
                    if exc.status_code == 404:
                        continue
                    raise
                if schedule.project.project_name == project_name:
                    project_ids.add(schedule.project.project_id)
        project_id = next(iter(project_ids)) if len(project_ids) == 1 else None
        workflow_targets.append(
            WorkflowProjectTarget(
                project_name=project_name,
                project_id=project_id,
                target_workflows=workflow_inventory.workflows.get(project_name, ()),
                target_schedule_ids=workflow_inventory.schedule_ids.get(project_name, ()),
                monitor_status=(
                    MonitorStatus.MONITOR
                    if project_id is not None
                    else MonitorStatus.NEEDS_REVIEW
                ),
            )
        )

    query_summaries = await saved_query_client.fetch_queries()
    queries_by_key: defaultdict[
        tuple[str, str, str],
        list[SavedQuerySummary],
    ] = defaultdict(list)
    for query in query_summaries:
        queries_by_key[
            _query_match_key(
                query.query_name,
                query.database.database_name,
                query.owner.owner_name,
            )
        ].append(query)

    saved_query_targets: list[SavedQueryTarget] = []
    for inventory_item in saved_query_inventory:
        matches = queries_by_key[
            _query_match_key(
                inventory_item.query_name,
                inventory_item.database,
                inventory_item.owner,
            )
        ]
        matched = matches[0] if len(matches) == 1 else None
        saved_query_targets.append(
            SavedQueryTarget(
                query_id=matched.query_id if matched is not None else None,
                query_name=(
                    matched.query_name
                    if matched is not None
                    else inventory_item.query_name
                ),
                database=inventory_item.database,
                owner=inventory_item.owner,
                monitor_status=(
                    MonitorStatus.MONITOR
                    if matched is not None
                    else MonitorStatus.NEEDS_REVIEW
                ),
            )
        )

    config = ResourceTargetsConfig(
        workflow_projects=tuple(workflow_targets),
        saved_queries=tuple(saved_query_targets),
    )
    return (
        config,
        TargetImportSummary(
            workflow_project_count=len(workflow_targets),
            workflow_needs_review_count=sum(
                target.monitor_status == MonitorStatus.NEEDS_REVIEW
                for target in workflow_targets
            ),
            saved_query_count=len(saved_query_targets),
            saved_query_needs_review_count=sum(
                target.monitor_status == MonitorStatus.NEEDS_REVIEW
                for target in saved_query_targets
            ),
        ),
    )


def resource_targets_to_yaml_bytes(config: ResourceTargetsConfig) -> bytes:
    """対象マスターを人がレビューできる決定的なYAMLへ変換する。

    引数:
        config: API照合後のWorkflow・登録クエリ対象。
    戻り値:
        IDと照合情報だけを含みSQLやAPIレスポンスを含まないUTF-8 YAML。
    """
    payload = {
        "workflow_projects": [
            {
                "project_name": target.project_name,
                "project_id": target.project_id,
                "target_workflows": list(target.target_workflows),
                "target_schedule_ids": list(target.target_schedule_ids),
                "monitor_status": target.monitor_status.value,
            }
            for target in config.workflow_projects
        ],
        "saved_queries": [
            {
                "query_id": target.query_id,
                "query_name": target.query_name,
                "database": target.database,
                "owner": target.owner,
                "monitor_status": target.monitor_status.value,
            }
            for target in config.saved_queries
        ],
    }
    content = yaml.safe_dump(
        payload,
        allow_unicode=True,
        sort_keys=False,
    )
    return content.encode("utf-8")


def write_resource_targets(
    path: Path,
    config: ResourceTargetsConfig,
) -> None:
    """対象マスターを同一ディレクトリの一時ファイル経由で置き換える。

    引数:
        path: 書き込み先`config/resource_targets.yml`。
        config: 保存する照合済み対象マスター。
    戻り値:
        なし。
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            dir=path.parent,
            prefix=f".{path.name}.",
            suffix=".tmp",
            delete=False,
        ) as temporary:
            temporary.write(resource_targets_to_yaml_bytes(config))
            temporary.flush()
            os.fsync(temporary.fileno())
            temporary_path = Path(temporary.name)
        os.replace(temporary_path, path)
        temporary_path = None
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def _read_workbook(
    workbook_path: Path,
) -> tuple[_WorkflowInventory, tuple[_SavedQueryInventoryItem, ...]]:
    """棚卸しExcelの3シートから利用中行だけを読み取る。

    引数:
        workbook_path: 読み取り対象の`.xlsx`ファイル。
    戻り値:
        project別Workflow・scheduleと登録クエリ照合キー。
    """
    if not workbook_path.is_file():
        raise ValueError(f"inventory workbook was not found: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        required_sheets = {
            _WORKFLOW_SHEET,
            _WORKFLOW_SCHEDULE_SHEET,
            _SAVED_QUERY_SHEET,
        }
        missing_sheets = required_sheets - set(workbook.sheetnames)
        if missing_sheets:
            raise ValueError(
                "inventory workbook did not include required sheets: "
                + ", ".join(sorted(missing_sheets))
            )

        workflow_rows = _active_rows(
            workbook[_WORKFLOW_SHEET],
            required_headers=("プロジェクト名", "ワークフロー名", "利用状態"),
        )
        schedule_rows = _active_rows(
            workbook[_WORKFLOW_SCHEDULE_SHEET],
            required_headers=(
                "プロジェクト名",
                "ワークフロー名",
                "スケジュールID",
                "利用状態",
            ),
        )
        query_rows = _active_rows(
            workbook[_SAVED_QUERY_SHEET],
            required_headers=("クエリ名", "データベース", "作成者", "利用状態"),
        )
    finally:
        workbook.close()

    workflows: defaultdict[str, set[str]] = defaultdict(set)
    schedule_ids: defaultdict[str, set[str]] = defaultdict(set)
    for row in workflow_rows:
        workflows[row["プロジェクト名"]].add(row["ワークフロー名"])
    for row in schedule_rows:
        project_name = row["プロジェクト名"]
        workflows[project_name].add(row["ワークフロー名"])
        schedule_id = row["スケジュールID"]
        if not schedule_id.isdigit():
            raise ValueError("inventory schedule ID must contain digits only")
        schedule_ids[project_name].add(schedule_id)

    saved_queries = tuple(
        _SavedQueryInventoryItem(
            query_name=row["クエリ名"],
            database=row["データベース"],
            owner=row["作成者"],
        )
        for row in query_rows
    )
    duplicate_query_keys = len(saved_queries) - len(
        {
            _query_match_key(item.query_name, item.database, item.owner)
            for item in saved_queries
        }
    )
    if duplicate_query_keys:
        raise ValueError("inventory contained duplicate active saved query keys")

    return (
        _WorkflowInventory(
            workflows={
                project: tuple(sorted(names))
                for project, names in sorted(workflows.items())
            },
            schedule_ids={
                project: tuple(sorted(ids, key=lambda item: (len(item), item)))
                for project, ids in sorted(schedule_ids.items())
            },
        ),
        saved_queries,
    )


def _active_rows(
    worksheet: Worksheet,
    *,
    required_headers: tuple[str, ...],
) -> tuple[dict[str, str], ...]:
    """1シートのheaderを確認し利用状態が利用中の行だけを返す。

    引数:
        worksheet: openpyxlの読み取り専用worksheet。
        required_headers: このシートで必須とする列名。
    戻り値:
        必須列を文字列へ正規化した利用中行。
    """
    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"inventory sheet was empty: {worksheet.title}") from exc
    headers = tuple(_cell_text(value) for value in raw_headers)
    missing_headers = set(required_headers) - set(headers)
    if missing_headers:
        raise ValueError(
            f"inventory sheet {worksheet.title} did not include headers: "
            + ", ".join(sorted(missing_headers))
        )
    indexes = {header: headers.index(header) for header in required_headers}

    active: list[dict[str, str]] = []
    for raw_row in rows:
        status = _row_value(raw_row, indexes["利用状態"])
        if status != _ACTIVE_VALUE:
            continue
        item = {
            header: _row_value(raw_row, indexes[header])
            for header in required_headers
        }
        empty = [header for header, value in item.items() if not value]
        if empty:
            raise ValueError(
                f"inventory active row in {worksheet.title} had blank fields: "
                + ", ".join(sorted(empty))
            )
        active.append(item)
    return tuple(active)


def _row_value(row: tuple[object, ...], index: int) -> str:
    """Excel行の指定列を前後空白なしの文字列へ変換する。

    引数:
        row: openpyxlが返した1行。
        index: 0始まりの列位置。
    戻り値:
        空cellなら空文字、それ以外は文字列表現。
    """
    if index >= len(row):
        return ""
    return _cell_text(row[index])


def _cell_text(value: object) -> str:
    """Excel cell値を照合用文字列へ正規化する。

    引数:
        value: openpyxlが返したcell値。
    戻り値:
        Noneなら空文字、数値IDは小数なし、その他は前後空白除去文字列。
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _query_match_key(
    query_name: str,
    database: str,
    owner: str,
) -> tuple[str, str, str]:
    """登録クエリのExcel行とAPI一覧を照合する正規化keyを作る。

    引数:
        query_name: 登録クエリ名。
        database: 実行対象database名。
        owner: 作成者・所有者表示名。
    戻り値:
        前後空白と大文字小文字の差を除いた3項目key。
    """
    return (
        query_name.strip().casefold(),
        database.strip().casefold(),
        owner.strip().casefold(),
    )
