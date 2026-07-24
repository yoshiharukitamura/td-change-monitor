from __future__ import annotations

import asyncio
from pathlib import Path

from openpyxl import Workbook

from td_change_monitor.config import MonitorStatus
from td_change_monitor.models import (
    SavedQueryDatabaseReference,
    SavedQueryOwnerReference,
    SavedQuerySummary,
    WorkflowDefinitionSummary,
    WorkflowProjectReference,
    WorkflowReference,
    WorkflowScheduleDetail,
)
from td_change_monitor.target_import import (
    build_resource_targets_from_workbook,
    resource_targets_to_yaml_bytes,
)


class FakeWorkflowLookup:
    """project名ごとに固定Workflow一覧を返すテスト用API。"""

    async def fetch_workflow_page_by_project_name(
        self,
        project_name: str,
        *,
        last_id: int = 0,
        count: int = 3,
    ) -> tuple[WorkflowDefinitionSummary, ...]:
        assert last_id == 0
        assert count == 3
        if project_name == "project_a":
            return (
                WorkflowDefinitionSummary(
                    workflow_id="10",
                    workflow_name="main",
                    project=WorkflowProjectReference(
                        project_id="100",
                        project_name="project_a",
                    ),
                    revision="revision",
                    timezone="UTC",
                ),
            )
        return ()

    async def fetch_schedule(self, schedule_id: str) -> WorkflowScheduleDetail:
        assert schedule_id == "200"
        return WorkflowScheduleDetail(
            schedule_id="200",
            project=WorkflowProjectReference(
                project_id="100",
                project_name="project_a",
            ),
            workflow=WorkflowReference(
                workflow_id="10",
                workflow_name="main",
            ),
            enabled=True,
        )


class FakeSavedQueryLookup:
    """Query ID付き一覧を返すテスト用API。"""

    async def fetch_queries(self) -> tuple[SavedQuerySummary, ...]:
        return (
            SavedQuerySummary(
                query_id="300",
                query_name=" query_a",
                database=SavedQueryDatabaseReference(
                    database_id="1",
                    database_name="db",
                ),
                owner=SavedQueryOwnerReference(
                    owner_id="2",
                    owner_name="Sample User",
                ),
                engine_type="trino",
                engine_version="stable",
                connector_config=None,
                cron=None,
                timezone="UTC",
                delay=0,
                priority=0,
                retry_limit=0,
                description=None,
                draft=False,
            ),
        )


def build_inventory_workbook(path: Path) -> None:
    """必要な3シートを持つ最小棚卸しExcelを作る。"""
    workbook = Workbook()
    workflow = workbook.active
    workflow.title = "登録Workflow一覧"
    workflow.append(["プロジェクト名", "ワークフロー名", "利用状態"])
    workflow.append(["project_a", "main", "利用中"])
    workflow.append(["project_b", "other", "利用中"])
    workflow.append(["ignored", "old", "利用なし"])

    schedule = workbook.create_sheet("Workflowスケジュール一覧")
    schedule.append(
        ["プロジェクト名", "ワークフロー名", "スケジュールID", "利用状態"]
    )
    schedule.append(["project_a", "main", 200, "利用中"])

    query = workbook.create_sheet("登録クエリ一覧")
    query.append(["クエリ名", "データベース", "作成者", "利用状態"])
    query.append(["query_a", "db", "Sample User", "利用中"])
    query.append(["missing", "db", "Sample User", "利用中"])
    workbook.save(path)


def test_import_matches_stable_ids_and_marks_unresolved_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "inventory.xlsx"
    build_inventory_workbook(workbook_path)

    config, summary = asyncio.run(
        build_resource_targets_from_workbook(
            workbook_path,
            workflow_client=FakeWorkflowLookup(),
            saved_query_client=FakeSavedQueryLookup(),
        )
    )

    assert config.workflow_projects[0].project_id == "100"
    assert config.workflow_projects[0].target_schedule_ids == ("200",)
    assert config.workflow_projects[1].monitor_status == MonitorStatus.NEEDS_REVIEW
    assert config.saved_queries[0].query_id == "300"
    assert config.saved_queries[0].query_name == " query_a"
    assert config.saved_queries[1].monitor_status == MonitorStatus.NEEDS_REVIEW
    assert summary.as_dict() == {
        "workflow_project_count": 2,
        "workflow_needs_review_count": 1,
        "saved_query_count": 2,
        "saved_query_needs_review_count": 1,
    }

    yaml_content = resource_targets_to_yaml_bytes(config)
    assert b"queryString" not in yaml_content
    assert b"connectorConfig" not in yaml_content
